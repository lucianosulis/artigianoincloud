from flask import (
    Blueprint, flash, g, redirect, render_template, request, url_for, current_app, jsonify, json, session
)
from flask_paginate import Pagination, get_page_args, get_page_parameter
from werkzeug.exceptions import abort

from flaskr.auth import login_required
from flaskr.db import get_db
from flask import send_file, send_from_directory

from datetime import datetime, timedelta
import calendar
from dateutil.easter import *
#import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment 
from openpyxl.utils import get_column_letter
#from openpyxl.styles import Font
#from openpyxl.styles import Alignment
import io

bp = Blueprint('report', __name__)

@bp.route('/report', methods=('GET',))
@login_required
def index():
    if g.role != "ADMIN" and g.role != "SEGRETERIA":
         error = 'Non sei autorizzato a questa funzione.'
         flash(error)
         return redirect(url_for("calendar.show_cal"))
    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute(
        'SELECT o.id AS order_id, CONCAT(DATE_FORMAT(o.date,"%d/%m/%y"), " - ", c.full_name, " - ", REPLACE(o.description,"""","")) AS order_desc'
        ' FROM p_order o ' 
        ' INNER JOIN customer c ON o.customer_id = c.id ' 
        ' WHERE o.closed=0'
        ' ORDER BY c.full_name ASC, o.date DESC'
    )
    orders=cursor.fetchall()
    anag_people = get_anag_people()

    return render_template('report/index.html', orders=orders, anag_people=anag_people)

@bp.route('/report/ore_lav_people', methods=['POST'])
def ore_lav_people():
    # Recupero i parametri dal form della card
    start = request.form.get('date1')
    end = request.form.get('date2')
    people_ids_list = request.form.getlist('people')
    people_ids = ",".join(people_ids_list)
    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute(
        'select c.full_name as Cliente, o.id as ID_Ordine, o.description as Attività, DATE_FORMAT(o.date,"%d/%m/%y") as Data_ordine, CONCAT(p.surname, " ",p.name) as Operatore, DATE_FORMAT(t.date,"%d/%m/%y") as Data_attività, t.ore_lav as Ore_lavorate ' 
        'from p_order o ' 
        'inner join timesheet t on o.id = t.order_id ' 
        'inner join people p on p.id = t.people_id '
        'inner join customer c on c.id = o.customer_id '
        'where  t.date >= %s and t.date <=%s and p.id in (%s) '
        'order by t.date ',(start,end,people_ids)
    )
    rows=cursor.fetchall()
    
    excel_file = generate_excel_response(rows,sheet_name="Ore lavorate")
    if not excel_file:
        flash("Nessun dato trovato")
        return redirect(url_for('report.index')) # Torna alla pagina dei report
    return send_file(
            excel_file,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="report_ore_lav_people.xlsx")


@bp.route('/report/order_activities', methods=['POST'])
def order_activities():
    # Recupero i parametri dal form della card
    #start = request.form.get('start')
    #end = request.form.get('end')
    order_id = request.form.get('order_id')
    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute(
        'SELECT a.title AS Titolo, c.full_name AS Cliente, DATE_FORMAT(a.start,"%d/%m/%y") AS Data_inizio, DATE_FORMAT(a.end,"%d/%m/%y") AS Data_fine, s.address AS Indirizzo, s.city AS Città '
        ' FROM activity a ' 
        ' INNER JOIN p_order o ON o.id = a.p_order_id '
        ' INNER JOIN site s ON s.id = a.site_id '
        ' INNER JOIN customer c ON o.customer_id = c.id ' 
        ' WHERE o.closed=0 AND a.p_order_id = %s'
        ' ORDER BY c.full_name ASC, o.date DESC',(order_id,)
    )
    rows=cursor.fetchall()
    excel_file = generate_excel_response(rows,sheet_name="Attività")
    if not excel_file:
        flash("Nessun dato trovato")
        return redirect(url_for('report.index')) # Torna alla pagina dei report
    return send_file(
            excel_file,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="report_attivita_ordine.xlsx")

@bp.route('/report/order_teams', methods=['POST'])
def order_teams():
    # Recupero i parametri dal form della card
    #start = request.form.get('start')
    #end = request.form.get('end')
    order_id = request.form.get('order_id')
    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute(
        'SELECT a.title AS Titolo, c.full_name AS Cliente, ' 
        ' DATE_FORMAT(a.start,"%d/%m/%y") AS Inizio_attività, ' 
        ' DATE_FORMAT(a.end,"%d/%m/%y") AS Fine_attività, s.address AS Indirizzo, s.city AS Città, '
        ' CONCAT(p.surname, " ",p.name) AS Operatore, DATE_FORMAT(ts.date,"%d/%m/%y") AS Data, '
        ' ts.ore_lav AS Ore_lavorate '
        ' FROM activity a ' 
        ' INNER JOIN p_order o ON o.id = a.p_order_id '
        ' INNER JOIN site s ON s.id = a.site_id '
        ' INNER JOIN customer c ON o.customer_id = c.id ' 
        ' INNER JOIN timesheet ts ON a.id = ts.act_id '
        ' INNER JOIN people p ON p.id = ts.people_id '
        ' WHERE o.closed=0 AND a.p_order_id = %s'
        ,(order_id,)
    )
    rows=cursor.fetchall()
    excel_file = generate_excel_response(rows,sheet_name="Attività")
    if not excel_file:
        flash("Nessun dato trovato")
        return redirect(url_for('report.index')) # Torna alla pagina dei report
    return send_file(
            excel_file,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="report_operatori_attivita.xlsx")

@bp.route('/report/activity_cost', methods=['POST'])
def activity_cost():
    # Recupero i parametri dal form della card
    start = request.form.get('date1')
    end = request.form.get('date2')
    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute(
        '''select a.id as id_attivita,a.title as titolo, a.start as inizio, a.end as fine, 
        o.id as id_ordine,o.description as desc_ordine,c.id as id_cliente, 
        c.full_name as nome_cliente,s.address as indirizzo_sito,s.city as città_sito,tg.code as tag, 
        'ore_uomo' as tipo,
        tm.ore_lav as ore_uomo, "" as ore_mezzo,"" as km_mezzo, "" as costo_materiale,CONCAT(p.surname," ",p.name) as risorsa,tm.date as data 
        from activity a 
        inner join p_order o on o.id = a.p_order_id 
        inner join customer c on c.id = o.customer_id  
        inner join site s on s.id = a.site_id 
        left join rel_tag_activity rta on rta.activity_id = a.id 
        left join tag tg on rta.tag_id  = tg.id 
        inner join timesheet tm on tm.act_id = a.id 
        left join people p on p.id = tm.people_id 
        where a.start >= %s and a.start <= %s
        union all
        select a.id as id_attivita,a.title as titolo, a.start as inizio, a.end as fine, 
        o.id as id_ordine,o.description as desc_ordine,c.id as id_cliente, 
        c.full_name as nome_cliente,s.address as indirizzo_sito,s.city as città_sito,tg.code as tag, 
        'ore_mezzo' as tipo,
        "" as ore_uomo, tu.ore_lav as ore_mezzo,"" as km_mezzo, "" as costo_materiale, CONCAT(t.brand," ",t.model," ",t.license_plate ) as risorsa,tu.date as data
        from activity a 
        inner join p_order o on o.id = a.p_order_id 
        inner join customer c on c.id = o.customer_id  
        inner join site s on s.id = a.site_id 
        left join rel_tag_activity rta on rta.activity_id = a.id 
        inner join tag tg on rta.tag_id  = tg.id 
        inner join tool_usage tu on tu.act_id = a.id 
        inner join tool t on tu.tool_id = t.id 
        where a.start >= %s and a.start <= %s and tu.ore_lav <> 0
        union all
        select a.id as id_attivita,a.title as titolo, a.start as inizio, a.end as fine, 
        o.id as id_ordine,o.description as desc_ordine,c.id as id_cliente, 
        c.full_name as nome_cliente,s.address as indirizzo_sito,s.city as città_sito,tg.code as tag, 
        'km_mezzo' as tipo,
        "" as ore_uomo, "" as ore_mezzo,tu.km as km_mezzo, "" as costo_materiale, CONCAT(t.brand," ",t.model," ",t.license_plate ) as risorsa,tu.date as data
        from activity a 
        inner join p_order o on o.id = a.p_order_id 
        inner join customer c on c.id = o.customer_id  
        inner join site s on s.id = a.site_id 
        left join rel_tag_activity rta on rta.activity_id = a.id 
        inner join tag tg on rta.tag_id  = tg.id 
        inner join tool_usage tu on tu.act_id = a.id 
        inner join tool t on tu.tool_id = t.id 
        where a.start >= %s and a.start <= %s and tu.km <> 0
        union all
        select a.id as id_attivita,a.title as titolo, a.start as inizio, a.end as fine, 
        o.id as id_ordine,o.description as desc_ordine,c.id as id_cliente, 
        c.full_name as nome_cliente,s.address as indirizzo_sito,s.city as città_sito,tg.code as tag, 
        'costo_materiale' as tipo,
        "" as ore_uomo, "" as ore_mezzo,"" as km_mezzo, mu.cost as costo_materiale, mu.description  as risorsa,'' as data
        from activity a 
        inner join p_order o on o.id = a.p_order_id 
        inner join customer c on c.id = o.customer_id  
        inner join site s on s.id = a.site_id 
        left join rel_tag_activity rta on rta.activity_id = a.id 
        inner join tag tg on rta.tag_id  = tg.id 
        inner join material_usage mu on mu.activity_id = a.id 
        where a.start >= %s and a.start <= %s
        order by id_attivita,tipo''',(start,end,start,end,start,end,start,end)
    )
    rows=cursor.fetchall()
    excel_file = generate_excel_response(rows,sheet_name="Attività")
    if not excel_file:
        flash("Nessun dato trovato")
        return redirect(url_for('report.index')) # Torna alla pagina dei report
    return send_file(
            excel_file,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="report_costi_attivita.xlsx")

@bp.route('/report/order_revenue', methods=['POST'])
def order_revenue():
    # Recupero i parametri dal form della card
    start = request.form.get('date1')
    end = request.form.get('date2')
    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute("""
        select o.id as order_id,o.description as order_desc,o.date as order_date,o.order_type, o.closed, 
        o.amount as order_amount,o.customer_id,c.full_name as customer_name,
        GROUP_CONCAT(t.code SEPARATOR ',') AS tag 
        from p_order o 
        inner join customer c on c.id = o.customer_id 
        left join rel_tag_order rto on rto.p_order_id = o.id 
        left join tag t on t.id = rto.tag_id 
        where o.date >= %s and o.date <= %s 
        GROUP BY o.id 
        order by o.date """,(start,end)
        )
    rows=cursor.fetchall()
    excel_file = generate_excel_response(rows,sheet_name="Attività")
    if not excel_file:
        flash("Nessun dato trovato")
        return redirect(url_for('report.index')) # Torna alla pagina dei report
    return send_file(
            excel_file,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="report_ricavi_ordine.xlsx")


def generate_excel_response(query_results, sheet_name="Report"):
    if not query_results:
        return None
    # Creo un nuovo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attività"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    #ws.row_dimensions[1].height = 25
    #ws.row_dimensions[2].height = 25
    #ws.column_dimensions['A'].width = 10  # Imposta la larghezza della colonna in caratteri
    #ws.column_dimensions['B'].width = 5 
    #ws.column_dimensions['AH'].width = 7
    # 3. Estrai le intestazioni dalle chiavi del primo dizionario
    headers = list(query_results[0].keys())
    ws.append(headers)
    # Styling opzionale: Intestazione in grassetto
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for record in query_results:
        row_data = [record.get(key) for key in headers]
        ws.append(row_data) 
    #Adatto la larghezza delle colonne al contenuto
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Ottiene la lettera (A, B, C...)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2
    # 5. Salva in un buffer di memoria invece che su disco
    output = io.BytesIO()
    wb.save(output)
    output.seek(0) #Riavvolge "il nastro" all'inizio
    return output
        
def get_anag_people():
    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute(
        'SELECT id, CONCAT(surname," ",name) AS name FROM people WHERE cessato=0 ORDER BY surname,name ASC'
    )
    anag_people=cursor.fetchall()
    return anag_people


