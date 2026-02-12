from flask import (
    Blueprint, flash, g, redirect, render_template, request, url_for, current_app, jsonify, json, session
)
from flask_paginate import Pagination, get_page_args, get_page_parameter
from werkzeug.exceptions import abort

from flaskr.auth import login_required
from flaskr.db import get_db
from flask import send_file, send_from_directory

from datetime import datetime, timedelta
import calendar
from dateutil.easter import *
#import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment 
from openpyxl.utils import get_column_letter
#from openpyxl.styles import Font
#from openpyxl.styles import Alignment

bp = Blueprint('export_timesheet', __name__)

@bp.route('/export_timesheet', methods=('GET', 'POST'))
@login_required
def index():
    session["activity_first_page"] = 'Y'
    if g.role != "ADMIN" and g.role != "SEGRETERIA":
         error = 'Non sei autorizzato a questa funzione.'
         flash(error)
         return redirect(url_for("calendar.show_cal"))
    report_type = "1"
    if request.method == 'POST': 
        report_type = request.form['input_report_type']
        current_app.config.from_file("config.json", load=json.load)
        path = current_app.config["WS_PATH"]
        input = request.form['month']
        error = None
        if (not input) or (not report_type):
            error = 'Compila il mese/anno.'
        if error is not None:
            flash(error)
        else:
            yearStr = input.split('-')[0]
            year = int(yearStr)
            monthStr = input.split('-')[1]
            month = int(monthStr)
            last_month_day = calendar.monthrange(year,month)[1]
            last_month_dayStr = str(last_month_day)
            start_date = yearStr + "-" + monthStr + "-" + "01"
            end_date = yearStr + "-" + monthStr + "-" + last_month_dayStr
            sheets = ['', 'GENNAIO', 'FEBBRAIO', 'MARZO', 'APRILE', 'MAGGIO', 'GIUGNO', 'LUGLIO', 'AGOSTO', 'SETTEMBRE', 'OTTOBRE', 'NOVEMBRE', 'DICEMBRE' ] 
            sheet_name = sheets[month]
            if report_type == "2":
                file_presenze = current_app.config["WS_FILE_PRESENZE_PAGHE"] + " " + yearStr + "-" + monthStr + ".xlsx"       #"RACCOLTA_ORE_PAGHE.xlsx"
            else:
                file_presenze = current_app.config["WS_FILE_PRESENZE_COMPLETO"] + " " + yearStr + "-" + monthStr + ".xlsx"    #"RACCOLTA_ORE_COMPLETO.xlsx"

            # Crea un nuovo workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.row_dimensions[1].height = 25
            ws.row_dimensions[2].height = 25
            ws.column_dimensions['A'].width = 10  # Imposta la larghezza della colonna in caratteri
            ws.column_dimensions['B'].width = 5 
            ws.column_dimensions['AH'].width = 7
            
            font1 = Font(name='Arial', size=16, bold=True, italic=True, color='000000')
            font2 = Font(name='Arial', size=10, bold=True, italic=True, color='000000')
            font3 = Font(name='Arial', size=10, bold=True, italic=False, color='000000')
            font4 = Font(name='Arial', size=8, bold=True, italic=False, color='FF0000')
            font5 = Font(name='Arial', size=8, bold=False, italic=False, color='000000')
            font6 = Font(name='Arial', size=9, bold=True, italic=False, color='000000')

            ws['A1'].font = font1
            ws['M1'].font = font2
            ws['X1'].font = font1
            ws['AF1'].font = font1
            ws['A2'].font = font2
            ws['B2'].font = font2
            ws['B2'].fill = PatternFill(start_color="c0c0c0", end_color="c0c0c0", fill_type="solid")
            ws['AH2'].font = font2
            ws['AH2'].fill = PatternFill(start_color="c0c0c0", end_color="c0c0c0", fill_type="solid")
            ws['A1'] = "PRONTOGIARDINI S.R.L."
            ws['M1'] = "PRESTAZIONE PER GIORNATA MESE"
            ws['X1'] = sheet_name
            ws['AF1'] = yearStr
            ws['A2'] = "COGNOME E NOME"
            ws['A2'].alignment = Alignment(wrap_text=True)
            ws['B2'] = "TIPO"
            ws['AH2'] = "TOTALI"
            day = 1
            while day <= last_month_day: 
                cell = ws.cell(row=2, column=day+2, value=day)
                cell.font = font3
                day = day + 1 
            i = 1
            while i <= 31:
                ws.column_dimensions[ws.cell(row=1, column=i+2).column_letter].width = 3.2
                i = i + 1 
            ws.merge_cells('A1:L1')
            ws.merge_cells('M1:W1')
            ws.merge_cells('X1:AE1')
            ws.merge_cells('AF1:AH1')
            
            db = get_db()
            cursor = db.cursor(dictionary=True)
            query = 'SELECT id, CONCAT(surname," ",name) AS p_name, type, gg_paga FROM people WHERE cessato = 0'
            if report_type == "2": #Report ufficio paghe
                query = query + ' AND (type = "D" OR type = "P") '  #Solo dipendenti full e a chiamata o part-time
            query = query + ' ORDER BY p_name'
            cursor.execute(query)
            peoples = cursor.fetchall()
            i = 3
            for people in peoples:
                cursor.execute(
                    ' SELECT substr(t.date, 9, 2) AS dayStr, p.id AS people_id, CONCAT(p.surname," ",p.name) AS p_name, '
                    ' t.date, GROUP_CONCAT(t.ore_lav SEPARATOR ";") AS ore, '
                    ' GROUP_CONCAT(a.short_code SEPARATOR ";") AS short_code, '
                    ' GROUP_CONCAT(t.night SEPARATOR ";") AS night '
                    ' FROM timesheet t '
                    ' INNER JOIN people p ON t.people_id = p.id'
                    ' INNER JOIN act_type a ON a.id  = t.act_type_id '
                    ' WHERE p.id = %s AND t.date >= %s AND t.date <= %s '
                    ' GROUP BY substr(t.date, 9, 2), p.id, CONCAT(p.surname," ",p.name), '
                    ' t.date  ',
                    (people['id'], start_date, end_date)
                )
                ts_records = cursor.fetchall()
                #print("len(ts_records): " + str(len(ts_records)))
                print("people type: " + people['type'])
                print("people type: " + people['p_name'])
                if len(ts_records) >  0:
                    cell = ws['A' + str(i)]
                    cell.value = people['p_name']
                    #print("p_type: " + people['p_type'])
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = font3
                    ws.merge_cells('A' + str(i) + ':A' + str(i+9))
                    
                    gg_paga = people['gg_paga']
                    if not gg_paga:
                        gg_paga = 0
                    print("gg_paga: " + str(gg_paga))
                    gg_ced_tot = 0

                    for ts_record in ts_records:
                        dayNum = int(ts_record['dayStr'])
                        print("dayNum: " +  ts_record['dayStr'])
                        gg_ced_tot = gg_ced_tot + 1
                        row = i
                        col = 2 + dayNum
                        ore_arr = ts_record['ore'].split(";")
                        code_arr = ts_record['short_code'].split(";")
                        night_arr = ts_record['night'].split(";")
                        week_day = calendar.weekday(year,month,dayNum)
                        if week_day == 4:   #Venerdì
                            ore_STD = 7
                        elif week_day == 5 or week_day == 6:  #Sabato e domenica
                            ore_STD = 0
                        else:               #Altri giorni
                            ore_STD = 8
                        #Indipendentemente dal giorno della settimana, se è festivo metto a zero le ore standard
                        if isHoliday(year,month,dayNum):
                            ore_STD = 0
                        ore_LAV = 0
                        ore_ORD_TOT = 0
                        ore_NOTT_TOT = 0
                        ore_STR = 0
                        ore_NOTT_STR = 0
                        ore_FE = 0
                        ore_PR = 0
                        ore_PNR = 0
                        ore_M = 0
                        ore_CISOA = 0
                        ore_CP = 0
            
                        k = 0
                        while k < len(code_arr):
                            if code_arr[k] == "LAV":
                                ore_LAV = ore_LAV + float(ore_arr[k])
                                if night_arr[k] == "1":
                                    ore_NOTT_TOT = ore_NOTT_TOT + float(ore_arr[k])
                                else:
                                    ore_ORD_TOT = ore_ORD_TOT + float(ore_arr[k])
                            if code_arr[k] == "FE":
                                ore_FE = ore_FE + float(ore_arr[k])
                            if code_arr[k] == "PR":
                                ore_PR = ore_PR + float(ore_arr[k])
                            if code_arr[k] == "PNR":
                                ore_PNR = ore_PNR + float(ore_arr[k])
                            if code_arr[k] == "M":
                                ore_M = ore_M + float(ore_arr[k])
                            if code_arr[k] == "CISOA":
                                ore_CISOA = ore_CISOA + float(ore_arr[k])
                            if code_arr[k] == "CP":
                                ore_CP = ore_CP + float(ore_arr[k])
                            k = k + 1
                        
                        ore_NOLAV = ore_FE + ore_PR + ore_PNR + ore_M + ore_CISOA + ore_CP
                        print("ore_NOLAV:")
                        print(ore_NOLAV)
                        ore_STD_NETTE = ore_STD - ore_NOLAV
                        #ore_STR = ore_LAV - ore_STD_NETTE 
                        ore_DIFF1 = ore_NOTT_TOT - ore_STD_NETTE
                        if ore_DIFF1 >= 0:
                            ore_ORD = 0
                            ore_STR = ore_ORD_TOT
                            ore_NOTT = ore_STD_NETTE
                            ore_NOTT_STR = ore_DIFF1
                        else:
                            ore_NOTT = ore_NOTT_TOT
                            ore_NOTT_STR = 0
                            ore_DIFF2 = ore_ORD_TOT + ore_DIFF1
                            if ore_DIFF2 >= 0:
                                ore_ORD = -ore_DIFF1
                                ore_STR = ore_DIFF2
                            else:
                                ore_ORD = ore_ORD_TOT
                                ore_STR = 0
                             
                        if people['type'] == "P" and gg_ced_tot > gg_paga and gg_paga > 0:
                                ore_STR = ore_STR + ore_ORD
                                ore_NOTT_STR = ore_NOTT_STR + ore_NOTT
                                ore_ORD = 0
                                ore_NOTT = 0 

                        if report_type == "2":  #Report ufficio paghe
                            ore_STR = 0
                            ore_NOTT_STR = 0
                            
                        if ore_ORD > 0:
                            c = ws.cell(row=row, column=col, value=ore_ORD)
                            c = ws.cell(row=row, column=2, value="ORD")
                            c.font = font6
                            c = ws.cell(row=row, column=34, value="=SUM(C"+str(row)+":AG"+str(row)+")")
                            c.font = font6
                        if ore_NOTT > 0:
                            c = ws.cell(row=row + 1, column=col, value=ore_NOTT)
                            c = ws.cell(row=row + 1, column=2, value="NOTT")
                            c.font = font6
                            c = ws.cell(row=row + 1, column=34, value='=SUM(C'+str(row+1)+':AG'+str(row+1)+')')
                            c.font = font6
                        if ore_STR > 0:
                            c = ws.cell(row=row + 2, column=col, value=ore_STR)
                            c = ws.cell(row=row + 2, column=2, value="STR")
                            c.font = font6
                            c = ws.cell(row=row + 2, column=34, value="=SUM(C"+str(row+2)+":AG"+str(row+2)+")")
                            c.font = font6
                        if ore_NOTT_STR > 0:
                            c = ws.cell(row=row + 3, column=col, value=ore_NOTT_STR)
                            c = ws.cell(row=row + 3, column=2, value="N-STR")
                            c.font = font6
                            c = ws.cell(row=row + 3, column=34, value="=SUM(C"+str(row+3)+":AG"+str(row+3)+")")
                            c.font = font6
                        if ore_FE > 0:
                            c = ws.cell(row=row + 4, column=col, value=ore_FE)
                            c = ws.cell(row=row + 4, column=2, value="FE")
                            c.font = font6
                            c = ws.cell(row=row + 4, column=34, value="=SUM(C"+str(row+4)+":AG"+str(row+4)+")")
                            c.font = font6
                        if ore_CISOA > 0:
                            c = ws.cell(row=row + 5, column=col, value=ore_CISOA) 
                            c = ws.cell(row=row + 5, column=2, value="CISOA") 
                            c.font = font6 
                            c = ws.cell(row=row + 5, column=34, value="=SUM(C"+str(row+5)+":AG"+str(row+5)+")")
                            c.font = font6
                        if ore_PR > 0:
                            c = ws.cell(row=row + 6, column=col, value=ore_PR)
                            c = ws.cell(row=row + 6, column=2, value="PR")
                            c.font = font6
                            c = ws.cell(row=row + 6, column=34, value="=SUM(C"+str(row+6)+":AG"+str(row+6)+")")
                            c.font = font6
                        if ore_PNR > 0:
                            c = ws.cell(row=row + 7, column=col, value=ore_PNR) 
                            c = ws.cell(row=row + 7, column=2, value="PNR")
                            c.font = font6
                            c = ws.cell(row=row + 7, column=34, value="=SUM(C"+str(row+7)+":AG"+str(row+7)+")")
                            c.font = font6
                        if ore_M > 0:
                            c = ws.cell(row=row + 8, column=col, value=ore_M) 
                            c = ws.cell(row=row + 8, column=2, value="M") 
                            c.font = font6
                            c = ws.cell(row=row + 8, column=34, value="=SUM(C"+str(row+8)+":AG"+str(row+8)+")")
                            c.font = font6
                        if ore_CP > 0:
                            c = ws.cell(row=row + 9, column=col, value=ore_CP) 
                            c = ws.cell(row=row + 9, column=2, value="CP") 
                            c.font = font6
                            c = ws.cell(row=row + 9, column=34, value="=SUM(C"+str(row+9)+":AG"+str(row+9)+")")
                            c.font = font6
                        
                        col_letter = get_column_letter(col)
                        ws.column_dimensions[col_letter].width = 4
                        
                        if people['type'] == "D" and week_day <= 4 and not isHoliday(year,month,dayNum):  #Controllo ore sui dipendenti
                            ore_check =  ore_ORD + ore_NOTT + ore_NOLAV
                            if ore_check != ore_STD:
                                for row_index in range(row, row + 9):
                                    c = ws.cell(row=row_index, column=col) 
                                    c.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") #rosso

                    for day in range(1, last_month_day + 1): 
                        flag = True
                        for row_index in range(row, row + 9):
                            c =  ws.cell(row=row_index, column=day + 2)
                            if c.value != None:
                                flag = False
                                #print("c.value: " + str(c.value) + " - " + str(flag))
                        #print("flag: " + str(flag))
                        week_day = calendar.weekday(year,month,day)
                        if flag: #Tutte le celle di un giorno vuote (per una persona)
                           if people['type'] == "D" and week_day <= 4 and not isHoliday(year,month,day):  #Controllo ore dipendenti
                                for row_index in range(row, row + 9):
                                    c = ws.cell(row=row_index, column=day + 2)
                                    c.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") #rosso    
                    
                    i = i + 10
            n = 3
            while n <= last_month_day + 2:
                dayNum = n - 2 
                m = 2
                while m < i:
                    week_day = calendar.weekday(year,month,dayNum)
                    c = ws.cell(row=m, column=n)
                    if week_day == 5:
                        c.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid") #giallo
                    if isHoliday(year,month,dayNum):
                        c.fill = PatternFill(start_color="ff00ff", end_color="ff00ff", fill_type="solid") #cremisi
                    if week_day == 6:
                        c.fill = PatternFill(start_color="ff6d6d", end_color="ff6d6d", fill_type="solid") #rosso chiaro
                    m = m + 1
                n = n + 1

            for row in ws['C3:AG' + str(i)]:
                for cell in row:
                    cell.font = font5
                    cell.number_format = '0.0'

            for row_index in range(3, i):
                ws.row_dimensions[row_index].height = 10

            i = i + 1
            c = ws.cell(row=i, column=1, value="LEGENDA:")
            c.font = font4
            c = ws.cell(row=i, column=2)
            c.fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid") #giallo
            ws.merge_cells('B' + str(i) + ':C' + str(i))
            c = ws.cell(row=i, column=4, value="SABATO")
            c.font = font4
            ws.merge_cells('D' + str(i) + ':G' + str(i))
            c = ws.cell(row=i, column=8)
            c.fill = PatternFill(start_color="ff6d6d", end_color="ff6d6d", fill_type="solid") #rosso chiaro
            ws.merge_cells('H' + str(i) + ':I' + str(i))
            c = ws.cell(row=i, column=10, value="DOMENICA")
            c.font = font4
            ws.merge_cells('J' + str(i) + ':M' + str(i))
            c = ws.cell(row=i, column=14)
            c.fill = PatternFill(start_color="ff00ff", end_color="ff00ff", fill_type="solid") #giallo
            ws.merge_cells('N' + str(i) + ':O' + str(i))
            c = ws.cell(row=i, column=16, value="FESTIVITA'")
            c.font = font4
            ws.merge_cells('P' + str(i) + ':S' + str(i))

            i = i + 1
            c = ws.cell(row=i, column=1, value="FE = FERIE")
            c.font = font4
            c = ws.cell(row=i, column=2, value="M = MALATTIA")
            c.font = font4
            ws.merge_cells('B' + str(i) + ':E' + str(i))
            c = ws.cell(row=i, column=6, value="MT = MATERNITA'")
            c.font = font4
            ws.merge_cells('F' + str(i) + ':K' + str(i))
            c = ws.cell(row=i, column=12, value="CM = CONGEDO MATRIMONIALE")
            c.font = font4
            ws.merge_cells('L' + str(i) + ':R' + str(i))
            c = ws.cell(row=i, column=19, value="PNR = PERMESSO NON RETRIBUITO")
            c.font = font4
            ws.merge_cells('S' + str(i) + ':AB' + str(i))
            c = ws.cell(row=i, column=29, value="CP = CONGEDO PARENTALE")
            c.font = font4
            ws.merge_cells('AC' + str(i) + ':AH' + str(i))

            i = i + 1
            c = ws.cell(row=i, column=1, value="I = INFORTUNIO")
            c.font = font4
            c = ws.cell(row=i, column=2, value="S = SOSPENSIONE")
            c.font = font4
            ws.merge_cells('B' + str(i) + ':E' + str(i))
            c = ws.cell(row=i, column=6, value="CA = CORSO APPRENDISTI")
            c.font = font4
            ws.merge_cells('F' + str(i) + ':K' + str(i))
            c = ws.cell(row=i, column=12, value="PS = PERMESSI STUDIO")
            c.font = font4
            ws.merge_cells('L' + str(i) + ':R' + str(i))
            c = ws.cell(row=i, column=19, value="PR = PERMESSO RETRIBUITO (R.O.L.)")
            c.font = font4
            ws.merge_cells('S' + str(i) + ':AB' + str(i))
            c = ws.cell(row=i, column=29, value="LM = LICENZA MATRIMONIALE")
            c.font = font4
            ws.merge_cells('AC' + str(i) + ':AH' + str(i))

            i = i + 1
            c = ws.cell(row=i, column=1, value="NOTE:")
            c.font = font3
            ws.row_dimensions[i].height = 30
            ws.merge_cells('A' + str(i) + ':AH' + str(i))

            # Definisci il bordo
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            # Imposta il reticolo per un'area specifica (ad esempio, da A1 a C3)
            for row in ws['A1:AH' + str(i)]:
                for cell in row:
                    cell.border = border

            wb.save(path + file_presenze)
            flash("Report aggiornato con successo.")
    print("report_type: " + report_type)
    return render_template('export_timesheet/index.html', report_type = report_type)
    
@bp.route('/export_timesheet/download', methods=('POST',))
@login_required
def download():
    if g.role != "ADMIN" and g.role != "SEGRETERIA":
         error = 'Non sei autorizzato a questa funzione.'
         flash(error)
         return redirect(url_for("calendar.show_cal"))
    report_type = request.form['input_report_type']
    print("input_report_type: " + report_type)
    input = request.form['month']
    error = None
    if (not input) or (not report_type):
        error = 'Compila il mese/anno.'
    if error is not None:
        flash(error)
    else:
        yearStr = input.split('-')[0]
        monthStr = input.split('-')[1]
    current_app.config.from_file("config.json", load=json.load)
    path = current_app.config["WS_PATH"]
    if report_type == "2":
        file_presenze = current_app.config["WS_FILE_PRESENZE_PAGHE"] + " " + yearStr + "-" + monthStr + ".xlsx"      #"RACCOLTA_ORE_PAGHE.xlsx"
    else:
        file_presenze = current_app.config["WS_FILE_PRESENZE_COMPLETO"] + " " + yearStr + "-" + monthStr + ".xlsx"    #"RACCOLTA_ORE_COMPLETO.xlsx"
    return send_from_directory(path, file_presenze, as_attachment=True)
        
@bp.route('/export_timesheet/lock_month', methods=('POST',))
@login_required
def lock_month():
    if g.role != "ADMIN" and g.role != "SEGRETERIA":
         error = 'Non sei autorizzato a questa funzione.'
         flash(error)
         return redirect(url_for("calendar.show_cal"))
    input = request.form['month']
    error = None
    if (not input):
        error = 'Compila il mese/anno.'
    if error is not None:
        flash(error)
    else:
        yearStr = input.split('-')[0]
        year = int(yearStr)
        monthStr = input.split('-')[1]
        month = int(monthStr)
        last_month_day = calendar.monthrange(year,month)[1]
        last_month_dayStr = str(last_month_day)
        start_date = yearStr + "-" + monthStr + "-" + "01"
        end_date = yearStr + "-" + monthStr + "-" + last_month_dayStr
        db = get_db()
        cursor = db.cursor(dictionary=True)
        cursor.execute('UPDATE timesheet SET locked = True WHERE date >= %s AND date <= %s', (start_date,end_date))
        db.commit()
        flash("Mese bloccato")
    return render_template('export_timesheet/index.html')

@bp.route('/export_timesheet/unlock_month', methods=('POST',))
@login_required
def unlock_month():
    if g.role != "ADMIN" and g.role != "SEGRETERIA":
         error = 'Non sei autorizzato a questa funzione.'
         flash(error)
         return redirect(url_for("calendar.show_cal"))
    input = request.form['month']
    error = None
    if (not input):
        error = 'Compila il mese/anno.'
    if error is not None:
        flash(error)
    else:
        yearStr = input.split('-')[0]
        year = int(yearStr)
        monthStr = input.split('-')[1]
        month = int(monthStr)
        last_month_day = calendar.monthrange(year,month)[1]
        last_month_dayStr = str(last_month_day)
        start_date = yearStr + "-" + monthStr + "-" + "01"
        end_date = yearStr + "-" + monthStr + "-" + last_month_dayStr
        db = get_db()
        cursor = db.cursor(dictionary=True)
        cursor.execute('UPDATE timesheet SET locked = False WHERE date >= %s AND date <= %s', (start_date,end_date))
        db.commit()
        flash("Mese sbloccato")
    return render_template('export_timesheet/index.html')

def isHoliday(year,month,dayNum):
    #Calcola tutte le festività civili italiane compresa Pasquetta
    #Include anche la festa patronale di Mapello (29 settembre) 
    easter_date = easter(year)
    easter_monday_date = easter_date + timedelta(days=1)
    easter_monday_dayNum = easter_monday_date.day
    easter_monday_Month = easter_monday_date.month

    if (month == 1 and dayNum == 1) or \
        (month == 1 and dayNum == 6) or \
        (month == 4 and dayNum == 25) or \
        (month == 5 and dayNum == 1) or \
        (month == 6 and dayNum == 2) or \
        (month == 8 and dayNum == 15) or \
        (month == 9 and dayNum == 29) or \
        (month == 11 and dayNum == 1) or \
        (month == 12 and dayNum == 8) or \
        (month == 12 and dayNum == 25) or \
        (month == 12 and dayNum == 26) or \
        (month == easter_monday_Month and dayNum == easter_monday_dayNum):
        isHoliday = True
    else:
        isHoliday = False
    
    #print("mese " + str(month) + " giorno " + str(dayNum))
    #print(str(isHoliday))
    return isHoliday


